from __future__ import annotations

import json
from pathlib import Path
from typing import Dict

import openpyxl

from .models import (
    DesignStandard,
    FrameInput,
    FrameOptimizationInput,
    OptimizationConstraints,
    SectionProperty,
    SteelGrade,
    StoreyInput,
    StoreyLoadInput,
)
from .utils import normalize_name, safe_float, safe_int


def load_section_database(path: str | Path) -> Dict[str, SectionProperty]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sections: Dict[str, SectionProperty] = {}

    for ws in wb.worksheets:
        headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        header_map = {
            normalize_name(header): idx + 1
            for idx, header in enumerate(headers)
            if header is not None
        }

        for row in range(5, ws.max_row + 1):
            raw_name = ws.cell(row, 1).value
            if raw_name is None:
                continue

            name = normalize_name(raw_name)
            common = dict(
                name=name,
                weight_kg_per_m=safe_float(ws.cell(row, header_map["WEIGHT"]).value),
                area_mm2=safe_float(ws.cell(row, header_map["AREA"]).value),
                inertia_mm4=safe_float(ws.cell(row, header_map["SECOND MOMENT OF AREA"]).value) * 1e6,
                elastic_modulus_mm3=safe_float(ws.cell(row, header_map["ELASTIC SECTION MODULUS"]).value) * 1e3,
            )
            maybe_section_class = ws.cell(row, ws.max_column).value
            if maybe_section_class not in (None, ""):
                common["section_class"] = safe_int(maybe_section_class)

            if ws.title == "I Section":
                section = SectionProperty(
                    shape="I Section",
                    depth_mm=safe_float(ws.cell(row, header_map["DEPTH"]).value),
                    width_mm=safe_float(ws.cell(row, header_map["WIDTH"]).value),
                    **common,
                )
            elif ws.title == "Circular Hollow Section":
                section = SectionProperty(
                    shape="Circular Hollow Section",
                    diameter_mm=safe_float(ws.cell(row, header_map["EXT. DIAMETER"]).value),
                    **common,
                )
            elif ws.title == "Square Hollow Section":
                section = SectionProperty(
                    shape="Square Hollow Section",
                    side_mm=safe_float(ws.cell(row, header_map["SIDE DIMENSION"]).value),
                    **common,
                )
            else:
                continue

            sections[name] = section

    return sections


def load_material_database(path: str | Path) -> Dict[str, SteelGrade]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    grades: Dict[str, SteelGrade] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        grade = normalize_name(row[0])
        grades[grade] = SteelGrade(
            grade=grade,
            yield_strength_mpa=safe_float(row[1]),
            cost_sgd_per_kg=safe_float(row[2]),
        )
    return grades


def load_design_standard_database(path: str | Path) -> Dict[str, DesignStandard]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    standards: Dict[str, DesignStandard] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        code = normalize_name(row[0])
        standards[code] = DesignStandard(code=code, alpha=safe_float(row[1]), beta=safe_float(row[2]))
    return standards


def load_module1_input_json(path: str | Path) -> FrameInput:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    frame = FrameInput(
        num_storeys=int(data["num_storeys"]),
        beam_span_m=float(data["beam_span_m"]),
        design_standard=data["design_standard"],
        storeys=[StoreyInput(**row) for row in data["storeys"]],
    )
    return frame


def load_module2_input_json(path: str | Path) -> FrameOptimizationInput:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    c = data.get("constraints", {})
    constraints = OptimizationConstraints(
        utilization_min=float(c.get("utilization_min", 0.0)),
        utilization_max=float(c.get("utilization_max", 1.0)),
        beam_same_groups=c.get("beam_same_groups", []),
        column_same_groups=c.get("column_same_groups", []),
        allowed_steel_grades=c.get("allowed_steel_grades", []),
        beam_allowed_shapes=c.get(
            "beam_allowed_shapes",
            ["I Section", "Circular Hollow Section", "Square Hollow Section"],
        ),
        column_allowed_shapes=c.get(
            "column_allowed_shapes",
            ["I Section", "Circular Hollow Section", "Square Hollow Section"],
        ),
        beam_max_section_class_by_storey={str(k): int(v) for k, v in c.get("beam_max_section_class_by_storey", {}).items()},
        column_max_section_class_by_storey={str(k): int(v) for k, v in c.get("column_max_section_class_by_storey", {}).items()},
    )
    return FrameOptimizationInput(
        num_storeys=int(data["num_storeys"]),
        beam_span_m=float(data["beam_span_m"]),
        design_standard=data["design_standard"],
        storeys=[StoreyLoadInput(**row) for row in data["storeys"]],
        constraints=constraints,
    )
