from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def save_json(data: Dict[str, Any], path: str | Path) -> None:
    Path(path).write_text(json.dumps(data, indent=2), encoding="utf-8")


def _apply_basic_style(ws, max_data_col: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1][:max_data_col]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_data_col):
        for cell in row:
            cell.alignment = center

    ws.freeze_panes = "A2"


def export_module1_excel(results: Dict[str, Any], output_path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Module1 Results"
    headers = [
        "Storey", "Height (m)", "Dead load (kN/m)", "Live load (kN/m)", "Design load w (kN/m)",
        "Beam section", "Beam grade", "Beam Mmax (kN·m)", "Beam stress (MPa)", "Beam utilization U",
        "Beam cost (SGD)", "Column section", "Column grade", "Column load P (kN)",
        "Column stress (MPa)", "Column utilization U", "Column total cost (SGD)", "Storey total cost (SGD)",
    ]
    ws.append(headers)
    for row in results["results_by_storey"]:
        ws.append([
            row["storey"], row["height_m"], row["dead_load_kn_per_m"], row["live_load_kn_per_m"], row["design_load_kn_per_m"],
            row["beam_section"], row["beam_grade"], row["beam_mmax_knm"], row["beam_sigma_mpa"], row["beam_utilization_ratio"], row["beam_cost_sgd"],
            row["column_section"], row["column_grade"], row["column_total_load_kn"], row["column_sigma_mpa"], row["column_utilization_ratio"],
            row["column_cost_total_sgd"], row["storey_total_cost_sgd"],
        ])
    ws["T1"] = "Design Standard"
    ws["U1"] = results["design_standard"]["code"]
    ws["T2"] = "Alpha"
    ws["U2"] = results["design_standard"]["alpha"]
    ws["T3"] = "Beta"
    ws["U3"] = results["design_standard"]["beta"]
    ws["T5"] = "Total Cost (SGD)"
    ws["U5"] = results["total_cost_sgd"]
    _apply_basic_style(ws, len(headers))
    for col_idx, width in {1: 10,2: 12,3: 18,4: 18,5: 18,6: 18,7: 12,8: 18,9: 18,10: 18,11: 16,12: 18,13: 12,14: 18,15: 18,16: 18,17: 20,18: 20,20: 18,21: 18}.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    wb.save(output_path)


def export_module2_excel(results: Dict[str, Any], output_path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    if results["feasible"]:
        ws.title = "Module2 Results"
        headers = [
            "Storey", "Height (m)", "Dead load (kN/m)", "Live load (kN/m)", "Design load w (kN/m)",
            "Beam section", "Beam shape", "Beam grade", "Beam section class", "Beam utilization U", "Beam cost (SGD)",
            "Column section", "Column shape", "Column grade", "Column section class", "Column utilization U",
            "Column total cost (SGD)", "Storey total cost (SGD)",
        ]
        ws.append(headers)
        for row in results["results_by_storey"]:
            ws.append([
                row["storey"], row["height_m"], row["dead_load_kn_per_m"], row["live_load_kn_per_m"], row["design_load_kn_per_m"],
                row["beam_section"], row["beam_shape"], row["beam_grade"], row["beam_section_class"], row["beam_utilization_ratio"], row["beam_cost_sgd"],
                row["column_section"], row["column_shape"], row["column_grade"], row["column_section_class"], row["column_utilization_ratio"],
                row["column_cost_total_sgd"], row["storey_total_cost_sgd"],
            ])
        ws["T1"] = "Feasible"
        ws["U1"] = "YES"
        ws["T2"] = "Design Standard"
        ws["U2"] = results["design_standard"]["code"]
        ws["T3"] = "Alpha"
        ws["U3"] = results["design_standard"]["alpha"]
        ws["T4"] = "Beta"
        ws["U4"] = results["design_standard"]["beta"]
        ws["T6"] = "Total Cost (SGD)"
        ws["U6"] = results["total_cost_sgd"]
        _apply_basic_style(ws, len(headers))
    else:
        ws.title = "Infeasible"
        ws["A1"] = "Module 2 result"
        ws["A2"] = "FEASIBLE"
        ws["B2"] = "NO"
        ws["A4"] = "The given constraints do not have a feasible solution."
        row_cursor = 6
        for report in results.get("infeasibility_reports", []):
            ws.cell(row=row_cursor, column=1, value=f"Member type: {report['member_type']}")
            ws.cell(row=row_cursor, column=2, value=f"Storey group: {report['group']}")
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value=report["message"])
            row_cursor += 2
    wb.save(output_path)
